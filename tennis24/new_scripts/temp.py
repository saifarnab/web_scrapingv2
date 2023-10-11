import openpyxl


def check_data_exists():
    filename = 'new_atp.xlsx'
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(row[3])


def column_to_index(col_str):
    """Converts Excel-style column letter(s) to numerical index."""
    col_str = col_str.upper()
    col_index = 0
    for char in col_str:
        col_index = col_index * 26 + (ord(char) - ord('A')) + 1
    return col_index - 1  # Adjust to be 0-based index


def remove_spaces_in_columns(filename, start_column, end_column):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Iterate through the rows and columns to remove spaces in the specified range
    for row in sheet.iter_rows():
        for col_idx in range(start_column, end_column + 1):
            cell = row[col_idx]
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace(' ', '')

    # Save the modified workbook
    new_filename = 'wtp_v2.xlsx'
    workbook.save(new_filename)
    print(f"Spaces removed and saved to {new_filename}")


# # Specify the Excel file, start column ("IR"), and end column ("LN")
# excel_file = 'new_wtp.xlsx'
# start_column = column_to_index("IR")
# end_column = column_to_index("LN")
#
# # Remove spaces in the specified columns
# remove_spaces_in_columns(excel_file, start_column, end_column)


def replace_slash_with_pipe(filename, column_index):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Iterate through the rows and replace '/' with '|'
    for row in sheet.iter_rows():
        cell = row[column_index]
        if cell.value and isinstance(cell.value, str):
            cell.value = cell.value.replace('/', '|')

    # Save the modified workbook
    new_filename = 'wtp_v3.xlsx'
    workbook.save(new_filename)
    print(f"Replaced '/' with '|' and saved to {new_filename}")


# Specify the Excel file and the column index (3rd column, 0-indexed)
excel_file = 'wtp_v2.xlsx'
column_index = 3  # Index 3 in 0-based indexing

# Replace '/' with '|' in the specified column
replace_slash_with_pipe(excel_file, column_index)
