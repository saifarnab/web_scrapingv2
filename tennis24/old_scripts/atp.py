import os
import socket
import time

import openpyxl
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


def config_driver() -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument(f'user-agent={UserAgent().random}')
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def create_directory_if_not_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def append_to_excel(data):
    filename = 'atp_4.xlsx'
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def get_wta_single_tournaments():
    v1 = [ 'https://www.tennis24.com/atp-singles/wimbledon/', 'https://www.tennis24.com/atp-singles/winston-salem/',
          'https://www.tennis24.com/atp-singles/zagreb/', 'https://www.tennis24.com/atp-singles/zaragoza/',
          'https://www.tennis24.com/atp-singles/zhuhai/']
    return v1


def create_excel_with_header():
    filename = 'atp_4.xlsx'
    if os.path.exists(filename) is True:
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        ['LINK', 'TOURNAMENT_NAME', 'SURFACE', 'STAGE', 'TOURNAMENT_TIME', 'PLAYER_1', 'PLAYER_2', 'RANK_1', 'RANK_2',
         'PLAYER_1_SET_WON', 'PLAYER_2_SET_WON', 'PLAYER_1_SET_1_GAMES_WON', 'PLAYER_2_SET_1_GAMES_WON',
         'PLAYER_1_SET_2_GAMES_WON', 'PLAYER_2_SET_2_GAMES_WON', 'PLAYER_1_SET_3_GAMES_WON',
         'PLAYER_2_SET_3_GAMES_WON',
         'PLAYER_1_SET_4_GAMES_WON', 'PLAYER_2_SET_4_GAMES_WON', 'PLAYER_1_SET_5_GAMES_WON',
         'PLAYER_2_SET_5_GAMES_WON',
         'PLAYER_1_BET365', 'PLAYER_2_BET365', 'match_player1_aces', 'match_player2_aces',
         'match_player1_double_faults', 'match_player2_double_faults',
         'match_player1_1st_server_per', 'match_player2_1st_server_per',
         'match_player1_2nd_server_per', 'match_player2_2nd_server_per',
         'match_player1_break_point_saved', 'match_player2_break_point_saved',
         'match_player1_1st_return_points_won', 'match_player2_1st_return_points_won',
         'match_player1_2nd_return_points_won', 'match_player2_2nd_return_points_won',
         'match_player1_break_points_converted', 'match_player2_break_points_converted',
         'match_player1_winners', 'match_player2_winners',
         'match_player1_unforced_errors', 'match_player2_unforced_errors',
         'match_player1_net_points_won', 'match_player2_net_points_won',
         'match_player1_max_points_in_row', 'match_player2_max_points_in_row',
         'match_player1_service_points_won', 'match_player2_service_points_won',
         'match_player1_return_points_won', 'match_player2_return_points_won',
         'match_player1_total_points_won', 'match_player2_total_points_won',
         'match_player1_max_games_in_row', 'match_player2_max_games_in_row',
         'match_player1_service_games_won', 'match_player2_service_games_won',
         'match_player1_return_games_won', 'match_player2_return_games_won',
         'match_player1_total_games_won', 'match_player2_total_games_won', 'set1_player1_aces',
         'set1_player2_aces',
         'set1_player1_double_faults', 'set1_player2_double_faults',
         'set1_player1_1st_server_per', 'set1_player2_1st_server_per',
         'set1_player1_2nd_server_per', 'set1_player2_2nd_server_per',
         'set1_player1_break_point_saved', 'set1_player2_break_point_saved',
         'set1_player1_1st_return_points_won', 'set1_player2_1st_return_points_won',
         'set1_player1_2nd_return_points_won', 'set1_player2_2nd_return_points_won',
         'set1_player1_break_points_converted', 'set1_player2_break_points_converted',
         'set1_player1_winners', 'set1_player2_winners',
         'set1_player1_unforced_errors', 'set1_player2_unforced_errors',
         'set1_player1_net_points_won', 'set1_player2_net_points_won',
         'set1_player1_max_points_in_row', 'set1_player2_max_points_in_row',
         'set1_player1_service_points_won', 'set1_player2_service_points_won',
         'set1_player1_return_points_won', 'set1_player2_return_points_won',
         'set1_player1_total_points_won', 'set1_player2_total_points_won',
         'set1_player1_max_games_in_row', 'set1_player2_max_games_in_row',
         'set1_player1_service_games_won', 'set1_player2_service_games_won',
         'set1_player1_return_games_won', 'set1_player2_return_games_won',
         'set1_player1_total_games_won', 'set1_player2_total_games_won', 'set2_player1_aces',
         'set2_player2_aces',
         'set2_player1_double_faults', 'set2_player2_double_faults',
         'set2_player1_1st_server_per', 'set2_player2_1st_server_per',
         'set2_player1_2nd_server_per', 'set2_player2_2nd_server_per',
         'set2_player1_break_point_saved', 'set2_player2_break_point_saved',
         'set2_player1_1st_return_points_won', 'set2_player2_1st_return_points_won',
         'set2_player1_2nd_return_points_won', 'set2_player2_2nd_return_points_won',
         'set2_player1_break_points_converted', 'set2_player2_break_points_converted',
         'set2_player1_winners', 'set2_player2_winners',
         'set2_player1_unforced_errors', 'set2_player2_unforced_errors',
         'set2_player1_net_points_won', 'set2_player2_net_points_won',
         'set2_player1_max_points_in_row', 'set2_player2_max_points_in_row',
         'set2_player1_service_points_won', 'set2_player2_service_points_won',
         'set2_player1_return_points_won', 'set2_player2_return_points_won',
         'set2_player1_total_points_won', 'set2_player2_total_points_won',
         'set2_player1_max_games_in_row', 'set2_player2_max_games_in_row',
         'set2_player1_service_games_won', 'set2_player2_service_games_won',
         'set2_player1_return_games_won', 'set2_player2_return_games_won',
         'set2_player1_total_games_won', 'set2_player2_total_games_won', 'set3_player1_aces',
         'set3_player2_aces',
         'set3_player1_double_faults', 'set3_player2_double_faults',
         'set3_player1_1st_server_per', 'set3_player2_1st_server_per',
         'set3_player1_2nd_server_per', 'set3_player2_2nd_server_per',
         'set3_player1_break_point_saved', 'set3_player2_break_point_saved',
         'set3_player1_1st_return_points_won', 'set3_player2_1st_return_points_won',
         'set3_player1_2nd_return_points_won', 'set3_player2_2nd_return_points_won',
         'set3_player1_break_points_converted', 'set3_player2_break_points_converted',
         'set3_player1_winners', 'set3_player2_winners',
         'set3_player1_unforced_errors', 'set3_player2_unforced_errors',
         'set3_player1_net_points_won', 'set3_player2_net_points_won',
         'set3_player1_max_points_in_row', 'set3_player2_max_points_in_row',
         'set3_player1_service_points_won', 'set3_player2_service_points_won',
         'set3_player1_return_points_won', 'set3_player2_return_points_won',
         'set3_player1_total_points_won', 'set3_player2_total_points_won',
         'set3_player1_max_games_in_row', 'set3_player2_max_games_in_row',
         'set3_player1_service_games_won', 'set3_player2_service_games_won',
         'set3_player1_return_games_won', 'set3_player2_return_games_won',
         'set3_player1_total_games_won', 'set3_player2_total_games_won', 'set4_player1_aces',
         'set4_player2_aces',
         'set4_player1_double_faults', 'set4_player2_double_faults',
         'set4_player1_1st_server_per', 'set4_player2_1st_server_per',
         'set4_player1_2nd_server_per', 'set4_player2_2nd_server_per',
         'set4_player1_break_point_saved', 'set4_player2_break_point_saved',
         'set4_player1_1st_return_points_won', 'set4_player2_1st_return_points_won',
         'set4_player1_2nd_return_points_won', 'set4_player2_2nd_return_points_won',
         'set4_player1_break_points_converted', 'set4_player2_break_points_converted',
         'set4_player1_winners', 'set4_player2_winners',
         'set4_player1_unforced_errors', 'set4_player2_unforced_errors',
         'set4_player1_net_points_won', 'set4_player2_net_points_won',
         'set4_player1_max_points_in_row', 'set4_player2_max_points_in_row',
         'set4_player1_service_points_won', 'set4_player2_service_points_won',
         'set4_player1_return_points_won', 'set4_player2_return_points_won',
         'set4_player1_total_points_won', 'set4_player2_total_points_won',
         'set4_player1_max_games_in_row', 'set4_player2_max_games_in_row',
         'set4_player1_service_games_won', 'set4_player2_service_games_won',
         'set4_player1_return_games_won', 'set4_player2_return_games_won',
         'set4_player1_total_games_won', 'set4_player2_total_games_won', 'set5_player1_aces',
         'set5_player2_aces',
         'set5_player1_double_faults', 'set5_player2_double_faults',
         'set5_player1_1st_server_per', 'set5_player2_1st_server_per',
         'set5_player1_2nd_server_per', 'set5_player2_2nd_server_per',
         'set5_player1_break_point_saved', 'set5_player2_break_point_saved',
         'set5_player1_1st_return_points_won', 'set5_player2_1st_return_points_won',
         'set5_player1_2nd_return_points_won', 'set5_player2_2nd_return_points_won',
         'set5_player1_break_points_converted', 'set5_player2_break_points_converted',
         'set5_player1_winners', 'set5_player2_winners',
         'set5_player1_unforced_errors', 'set5_player2_unforced_errors',
         'set5_player1_net_points_won', 'set5_player2_net_points_won',
         'set5_player1_max_points_in_row', 'set5_player2_max_points_in_row',
         'set5_player1_service_points_won', 'set5_player2_service_points_won',
         'set5_player1_return_points_won', 'set5_player2_return_points_won',
         'set5_player1_total_points_won', 'set5_player2_total_points_won',
         'set5_player1_max_games_in_row', 'set5_player2_max_games_in_row',
         'set5_player1_service_games_won', 'set5_player2_service_games_won',
         'set5_player1_return_games_won', 'set5_player2_return_games_won',
         'set5_player1_total_games_won', 'set5_player2_total_games_won', 'pbp_set_1_game_1',
         'pbp_set_1_game_2', 'pbp_set_1_game_3', 'pbp_set_1_game_4', 'pbp_set_1_game_5',
         'pbp_set_1_game_6', 'pbp_set_1_game_7', 'pbp_set_1_game_8', 'pbp_set_1_game_9', 'pbp_set_1_game_10',
         'pbp_set_1_game_11', 'pbp_set_1_game_12', 'pbp_set_1_game_13', 'pbp_set_1_game_14',
         'pbp_set_1_game_15',
         'pbp_set_1_game_1_serve', 'pbp_set_1_game_2_serve', 'pbp_set_1_game_3_serve',
         'pbp_set_1_game_4_serve',
         'pbp_set_1_game_5_serve', 'pbp_set_1_game_6_serve', 'pbp_set_1_game_7_serve',
         'pbp_set_1_game_8_serve',
         'pbp_set_1_game_9_serve', 'pbp_set_1_game_10_serve', 'pbp_set_1_game_11_serve',
         'pbp_set_1_game_12_serve',
         'pbp_set_1_game_13_serve', 'pbp_set_1_game_14_serve', 'pbp_set_1_game_15_serve', 'pbp_set_2_game_1',
         'pbp_set_2_game_2', 'pbp_set_2_game_3', 'pbp_set_2_game_4', 'pbp_set_2_game_5',
         'pbp_set_2_game_6', 'pbp_set_2_game_7', 'pbp_set_2_game_8', 'pbp_set_2_game_9', 'pbp_set_2_game_10',
         'pbp_set_2_game_11', 'pbp_set_2_game_12', 'pbp_set_2_game_13', 'pbp_set_2_game_14',
         'pbp_set_2_game_15',
         'pbp_set_2_game_1_serve', 'pbp_set_2_game_2_serve', 'pbp_set_2_game_3_serve',
         'pbp_set_2_game_4_serve',
         'pbp_set_2_game_5_serve', 'pbp_set_2_game_6_serve', 'pbp_set_2_game_7_serve',
         'pbp_set_2_game_8_serve',
         'pbp_set_2_game_9_serve', 'pbp_set_2_game_10_serve', 'pbp_set_2_game_11_serve',
         'pbp_set_2_game_12_serve',
         'pbp_set_2_game_13_serve', 'pbp_set_2_game_14_serve', 'pbp_set_2_game_15_serve', 'pbp_set_3_game_1',
         'pbp_set_3_game_2', 'pbp_set_3_game_3', 'pbp_set_3_game_4', 'pbp_set_3_game_5',
         'pbp_set_3_game_6', 'pbp_set_3_game_7', 'pbp_set_3_game_8', 'pbp_set_3_game_9', 'pbp_set_3_game_10',
         'pbp_set_3_game_11', 'pbp_set_3_game_12', 'pbp_set_3_game_13', 'pbp_set_3_game_14',
         'pbp_set_3_game_15',
         'pbp_set_3_game_1_serve', 'pbp_set_3_game_2_serve', 'pbp_set_3_game_3_serve',
         'pbp_set_3_game_4_serve',
         'pbp_set_3_game_5_serve', 'pbp_set_3_game_6_serve', 'pbp_set_3_game_7_serve',
         'pbp_set_3_game_8_serve',
         'pbp_set_3_game_9_serve', 'pbp_set_3_game_10_serve', 'pbp_set_3_game_11_serve',
         'pbp_set_3_game_12_serve',
         'pbp_set_3_game_13_serve', 'pbp_set_3_game_14_serve', 'pbp_set_3_game_15_serve', 'pbp_set_4_game_1',
         'pbp_set_4_game_2', 'pbp_set_4_game_3', 'pbp_set_4_game_4', 'pbp_set_4_game_5',
         'pbp_set_4_game_6', 'pbp_set_4_game_7', 'pbp_set_4_game_8', 'pbp_set_4_game_9', 'pbp_set_4_game_10',
         'pbp_set_4_game_11', 'pbp_set_4_game_12', 'pbp_set_4_game_13', 'pbp_set_4_game_14',
         'pbp_set_4_game_15',
         'pbp_set_4_game_1_serve', 'pbp_set_4_game_2_serve', 'pbp_set_4_game_3_serve',
         'pbp_set_4_game_4_serve',
         'pbp_set_4_game_5_serve', 'pbp_set_4_game_6_serve', 'pbp_set_4_game_7_serve',
         'pbp_set_4_game_8_serve',
         'pbp_set_4_game_9_serve', 'pbp_set_4_game_10_serve', 'pbp_set_4_game_11_serve',
         'pbp_set_4_game_12_serve',
         'pbp_set_4_game_13_serve', 'pbp_set_4_game_14_serve', 'pbp_set_4_game_15_serve', 'pbp_set_5_game_1',
         'pbp_set_5_game_2', 'pbp_set_5_game_3', 'pbp_set_5_game_4', 'pbp_set_5_game_5',
         'pbp_set_5_game_6', 'pbp_set_5_game_7', 'pbp_set_5_game_8', 'pbp_set_5_game_9', 'pbp_set_5_game_10',
         'pbp_set_5_game_11', 'pbp_set_5_game_12', 'pbp_set_5_game_13', 'pbp_set_5_game_14',
         'pbp_set_5_game_15',
         'pbp_set_5_game_1_serve', 'pbp_set_5_game_2_serve', 'pbp_set_5_game_3_serve',
         'pbp_set_5_game_4_serve',
         'pbp_set_5_game_5_serve', 'pbp_set_5_game_6_serve', 'pbp_set_5_game_7_serve',
         'pbp_set_5_game_8_serve',
         'pbp_set_5_game_9_serve', 'pbp_set_5_game_10_serve', 'pbp_set_5_game_11_serve',
         'pbp_set_5_game_12_serve',
         'pbp_set_5_game_13_serve', 'pbp_set_5_game_14_serve', 'pbp_set_5_game_15_serve'])
    workbook.save(filename)


def check_data_exists(name: str) -> bool:
    filename = 'atp_4.xlsx'
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row[0] is None:
            return False
        if name.strip() == row[0].strip():
            return True  # Data exists in the sheet
    return False  # Data does not exist in the sheet


def is_internet_available():
    try:
        # Try to connect to a well-known website
        socket.create_connection(("www.google.com", 80))
        return True
    except socket.error:
        return False


def wait_for_internet():
    print("Waiting for internet connection...")
    while not is_internet_available():
        time.sleep(5)  # Wait for 5 seconds before checking again

    print("Internet connection is back!")


def iterate_tournament(driver: webdriver.Chrome):
    links = get_wta_single_tournaments()
    counter = 1
    for link in links:
        print('---------------------------------------------')
        print(f'Tournament link = {link}')
        if not is_internet_available():
            wait_for_internet()
        driver.get(f'{link}/results/')

        year = int(driver.find_element(By.XPATH, '//div[@class="heading__info"]').text)
        if year < 2021:
            continue

        while True:
            more_results = driver.find_elements(By.XPATH, '//a[@class="event__more event__more--static"]')
            if len(more_results) > 0:
                driver.execute_script("arguments[0].click();", more_results[0])
                time.sleep(3)
            else:
                break

        match_links = []
        matches = driver.find_elements(By.XPATH, '//div[@title="Click for match detail!"]')
        for match in matches:
            match_id = match.get_attribute('id').split('g_2_')[-1]
            match_links.append(f'https://www.tennis24.com/match/{match_id}/#/match-summary/match-summary')

        print(f'Total matches = {len(match_links)}')

        for index, match_link in enumerate(match_links):

            if link.strip() == 'https://www.tennis24.com/atp-singles/wimbledon/' and index < 228:
                print('Data exists')
                continue

            # if check_data_exists(match_link):
            #     print(f'{match_link} already available ')
            #     continue
            start_time = time.time()

            if not is_internet_available():
                wait_for_internet()

            driver.get(match_link)
            time.sleep(3)
            header = driver.find_element(By.XPATH, '//span[@class="tournamentHeader__country"]').text

            if ',' in header:
                header = driver.find_element(By.XPATH, '//span[@class="tournamentHeader__country"]').text.split(',')
                tournament_name, surface_stage = header[0], header[1].split('-')
                try:
                    surface = surface_stage[0]
                    stage = surface_stage[1]
                except:
                    surface, stage = '', ''
            else:
                tournament_name, surface, stage = header, '', ''

            tournament_time = driver.find_element(By.XPATH, '//div[@class="duelParticipant__startTime"]//div').text
            try:
                players = driver.find_elements(By.XPATH,
                                               '//a[@class="participant__participantName participant__overflow "]')
                player1 = players[0].text
                player2 = players[1].text
            except:
                player1 = ''
                player2 = ''
            ranks = driver.find_elements(By.XPATH, '//div[@class="participant__participantRank"]')
            try:
                rank1 = ranks[0].text.replace('ATP', '').replace('\n', '').replace(':', '').replace('.', '').strip()
            except:
                rank1 = ''

            try:
                rank2 = ranks[1].text.replace('ATP', '').replace('\n', '').replace(':', '').replace('.', '').strip()
            except:
                rank2 = ''

            sets = driver.find_element(By.XPATH, '//div[@class="detailScore__wrapper"]').text.replace('\n', '').split(
                '-')

            try:
                player1_set_won = sets[0]
                player2_set_won = sets[1]
            except:
                player1_set_won = ''
                player2_set_won = ''

            try:
                player1_set1_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__home smh__part--1"]').text
                player2_set1_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__away smh__part--1"]').text
            except:
                player1_set1_games_won, player2_set1_games_won = '', ''

            try:
                player1_set2_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__home smh__part--2"]').text
                player2_set2_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__away smh__part--2"]').text
            except:
                player1_set2_games_won, player2_set2_games_won = '', ''

            try:
                player1_set3_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__home smh__part--3"]').text
                player2_set3_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__away smh__part--3"]').text
            except:
                player1_set3_games_won, player2_set3_games_won = '', ''

            try:
                player1_set4_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__home smh__part--4"]').text
                player2_set4_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__away smh__part--4"]').text
            except:
                player1_set4_games_won, player2_set4_games_won = '', ''

            try:
                player1_set5_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__home smh__part--5"]').text
                player2_set5_games_won = driver.find_element(By.XPATH,
                                                             '//div[@class="smh__part  smh__away smh__part--5"]').text
            except:
                player1_set5_games_won, player2_set5_games_won = '', ''

            try:
                WebDriverWait(driver, 3).until(
                    EC.visibility_of_element_located((By.XPATH, '//span[@class="oddsValueInner"]')))
                bet365s = driver.find_elements(By.XPATH, '//span[@class="oddsValueInner"]')
                player1_bet365 = bet365s[0].text
                player2_bet365 = bet365s[1].text

            except:
                player1_bet365, player2_bet365 = '', ''

            # match

            (match_player1_aces, match_player2_aces, match_player1_double_faults, match_player2_double_faults,
             match_player1_1st_server_per,
             match_player2_1st_server_per, match_player1_2nd_server_per, match_player2_2nd_server_per,
             match_player1_break_point_saved,
             match_player2_break_point_saved, match_player1_1st_return_points_won,
             match_player2_1st_return_points_won,
             match_player1_2nd_return_points_won, match_player2_2nd_return_points_won,
             match_player1_break_points_converted,
             match_player2_break_points_converted, match_player1_winners, match_player2_winners,
             match_player1_unforced_errors,
             match_player2_unforced_errors, match_player1_net_points_won, match_player2_net_points_won,
             match_player1_max_points_in_row,
             match_player2_max_points_in_row, match_player1_service_points_won, match_player2_service_points_won,
             match_player1_return_points_won, match_player2_return_points_won, match_player1_total_points_won,
             match_player2_total_points_won,
             match_player1_max_games_in_row, match_player2_max_games_in_row, match_player1_service_games_won,
             match_player2_service_games_won,
             match_player1_return_games_won, match_player2_return_games_won, match_player1_total_games_won,
             match_player2_total_games_won) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')
            # set1
            (set1_player1_aces, set1_player2_aces, set1_player1_double_faults, set1_player2_double_faults,
             set1_player1_1st_server_per,
             set1_player2_1st_server_per, set1_player1_2nd_server_per, set1_player2_2nd_server_per,
             set1_player1_break_point_saved,
             set1_player2_break_point_saved, set1_player1_1st_return_points_won, set1_player2_1st_return_points_won,
             set1_player1_2nd_return_points_won, set1_player2_2nd_return_points_won,
             set1_player1_break_points_converted,
             set1_player2_break_points_converted, set1_player1_winners, set1_player2_winners,
             set1_player1_unforced_errors,
             set1_player2_unforced_errors, set1_player1_net_points_won, set1_player2_net_points_won,
             set1_player1_max_points_in_row,
             set1_player2_max_points_in_row, set1_player1_service_points_won, set1_player2_service_points_won,
             set1_player1_return_points_won, set1_player2_return_points_won, set1_player1_total_points_won,
             set1_player2_total_points_won,
             set1_player1_max_games_in_row, set1_player2_max_games_in_row, set1_player1_service_games_won,
             set1_player2_service_games_won,
             set1_player1_return_games_won, set1_player2_return_games_won, set1_player1_total_games_won,
             set1_player2_total_games_won) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')

            # set2
            (set2_player1_aces, set2_player2_aces, set2_player1_double_faults, set2_player2_double_faults,
             set2_player1_1st_server_per,
             set2_player2_1st_server_per, set2_player1_2nd_server_per, set2_player2_2nd_server_per,
             set2_player1_break_point_saved,
             set2_player2_break_point_saved, set2_player1_1st_return_points_won, set2_player2_1st_return_points_won,
             set2_player1_2nd_return_points_won, set2_player2_2nd_return_points_won,
             set2_player1_break_points_converted,
             set2_player2_break_points_converted, set2_player1_winners, set2_player2_winners,
             set2_player1_unforced_errors,
             set2_player2_unforced_errors, set2_player1_net_points_won, set2_player2_net_points_won,
             set2_player1_max_points_in_row,
             set2_player2_max_points_in_row, set2_player1_service_points_won, set2_player2_service_points_won,
             set2_player1_return_points_won, set2_player2_return_points_won, set2_player1_total_points_won,
             set2_player2_total_points_won,
             set2_player1_max_games_in_row, set2_player2_max_games_in_row, set2_player1_service_games_won,
             set2_player2_service_games_won,
             set2_player1_return_games_won, set2_player2_return_games_won, set2_player1_total_games_won,
             set2_player2_total_games_won) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')

            # set3
            (set3_player1_aces, set3_player2_aces, set3_player1_double_faults, set3_player2_double_faults,
             set3_player1_1st_server_per,
             set3_player2_1st_server_per, set3_player1_2nd_server_per, set3_player2_2nd_server_per,
             set3_player1_break_point_saved,
             set3_player2_break_point_saved, set3_player1_1st_return_points_won, set3_player2_1st_return_points_won,
             set3_player1_2nd_return_points_won, set3_player2_2nd_return_points_won,
             set3_player1_break_points_converted,
             set3_player2_break_points_converted, set3_player1_winners, set3_player2_winners,
             set3_player1_unforced_errors,
             set3_player2_unforced_errors, set3_player1_net_points_won, set3_player2_net_points_won,
             set3_player1_max_points_in_row,
             set3_player2_max_points_in_row, set3_player1_service_points_won, set3_player2_service_points_won,
             set3_player1_return_points_won, set3_player2_return_points_won, set3_player1_total_points_won,
             set3_player2_total_points_won,
             set3_player1_max_games_in_row, set3_player2_max_games_in_row, set3_player1_service_games_won,
             set3_player2_service_games_won,
             set3_player1_return_games_won, set3_player2_return_games_won, set3_player1_total_games_won,
             set3_player2_total_games_won) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')

            # set4
            (set4_player1_aces, set4_player2_aces, set4_player1_double_faults, set4_player2_double_faults,
             set4_player1_1st_server_per,
             set4_player2_1st_server_per, set4_player1_2nd_server_per, set4_player2_2nd_server_per,
             set4_player1_break_point_saved,
             set4_player2_break_point_saved, set4_player1_1st_return_points_won, set4_player2_1st_return_points_won,
             set4_player1_2nd_return_points_won, set4_player2_2nd_return_points_won,
             set4_player1_break_points_converted,
             set4_player2_break_points_converted, set4_player1_winners, set4_player2_winners,
             set4_player1_unforced_errors,
             set4_player2_unforced_errors, set4_player1_net_points_won, set4_player2_net_points_won,
             set4_player1_max_points_in_row,
             set4_player2_max_points_in_row, set4_player1_service_points_won, set4_player2_service_points_won,
             set4_player1_return_points_won, set4_player2_return_points_won, set4_player1_total_points_won,
             set4_player2_total_points_won,
             set4_player1_max_games_in_row, set4_player2_max_games_in_row, set4_player1_service_games_won,
             set4_player2_service_games_won,
             set4_player1_return_games_won, set4_player2_return_games_won, set4_player1_total_games_won,
             set4_player2_total_games_won) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')

            # set5
            (set5_player1_aces, set5_player2_aces, set5_player1_double_faults, set5_player2_double_faults,
             set5_player1_1st_server_per,
             set5_player2_1st_server_per, set5_player1_2nd_server_per, set5_player2_2nd_server_per,
             set5_player1_break_point_saved,
             set5_player2_break_point_saved, set5_player1_1st_return_points_won, set5_player2_1st_return_points_won,
             set5_player1_2nd_return_points_won, set5_player2_2nd_return_points_won,
             set5_player1_break_points_converted,
             set5_player2_break_points_converted, set5_player1_winners, set5_player2_winners,
             set5_player1_unforced_errors,
             set5_player2_unforced_errors, set5_player1_net_points_won, set5_player2_net_points_won,
             set5_player1_max_points_in_row,
             set5_player2_max_points_in_row, set5_player1_service_points_won, set5_player2_service_points_won,
             set5_player1_return_points_won, set5_player2_return_points_won, set5_player1_total_points_won,
             set5_player2_total_points_won,
             set5_player1_max_games_in_row, set5_player2_max_games_in_row, set5_player1_service_games_won,
             set5_player2_service_games_won,
             set5_player1_return_games_won, set5_player2_return_games_won, set5_player1_total_games_won,
             set5_player2_total_games_won) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')

            starts = driver.find_elements(By.XPATH, "//a[@href='#/match-summary/match-statistics']")
            if len(starts) != 0:
                driver.execute_script("arguments[0].click();", starts[0])
                WebDriverWait(driver, 3).until(
                    EC.visibility_of_element_located(
                        (By.XPATH, '//div[@class="subFilter detail__subFilter detail__subFilter--stats"]')))

                match_elements = driver.find_elements(By.XPATH, '//div[@data-testid="wcl-statistics"]')
                for match_element in match_elements:
                    temp_data = match_element.text.replace('\n', '#').split('#')
                    if temp_data[1].strip() == 'Aces':
                        match_player1_aces, match_player2_aces = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Double Faults':
                        match_player1_double_faults, match_player2_double_faults = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == '1st Serve Percentage':
                        match_player1_1st_server_per, match_player2_1st_server_per = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == '2nd Serve Points Won':
                        match_player1_2nd_server_per, match_player2_2nd_server_per = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Break Points Saved':
                        match_player1_break_point_saved, match_player2_break_point_saved = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == '1st Return Points Won':
                        match_player1_1st_return_points_won, match_player2_1st_return_points_won = temp_data[0], \
                            temp_data[2]
                    elif temp_data[1].strip() == '2nd Return Points Won':
                        match_player1_2nd_return_points_won, match_player2_2nd_return_points_won = temp_data[0], \
                            temp_data[2]
                    elif temp_data[1].strip() == 'Break Points Converted':
                        match_player1_break_points_converted, match_player2_break_points_converted = temp_data[0], \
                            temp_data[2]
                    elif temp_data[1].strip() == 'Winners':
                        match_player1_winners, match_player2_winners = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Unforced Errors':
                        match_player1_unforced_errors, match_player2_unforced_errors = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Net Points Won':
                        match_player1_net_points_won, match_player2_net_points_won = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Max Points In Row':
                        match_player1_max_points_in_row, match_player2_max_points_in_row = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Service Points Won':
                        match_player1_service_points_won, match_player2_service_points_won = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Return Points Won':
                        match_player1_return_points_won, match_player2_return_points_won = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Total Points Won':
                        match_player1_total_points_won, match_player2_total_points_won = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Max Games In Row':
                        match_player1_max_games_in_row, match_player2_max_games_in_row = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Service Games Won':
                        match_player1_service_games_won, match_player2_service_games_won = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Return Games Won':
                        match_player1_return_games_won, match_player2_return_games_won = temp_data[0], temp_data[2]
                    elif temp_data[1].strip() == 'Total Games Won':
                        match_player1_total_games_won, match_player2_total_games_won = temp_data[0], temp_data[2]

                set_1_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 1"]')
                if len(set_1_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_1_clicks_elm[0])
                    time.sleep(1)
                    set1_elements = driver.find_elements(By.XPATH, '//div[@data-testid="wcl-statistics"]')
                    for set1_element in set1_elements:
                        temp_data = set1_element.text.replace('\n', '#').split('#')
                        if temp_data[1].strip() == 'Aces':
                            set1_player1_aces, set1_player2_aces = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Double Faults':
                            set1_player1_double_faults, set1_player2_double_faults = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Serve Percentage':
                            set1_player1_1st_server_per, set1_player2_1st_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '2nd Serve Points Won':
                            set1_player1_2nd_server_per, set1_player2_2nd_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Saved':
                            set1_player1_break_point_saved, set1_player2_break_point_saved = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Return Points Won':
                            set1_player1_1st_return_points_won, set1_player2_1st_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == '2nd Return Points Won':
                            set1_player1_2nd_return_points_won, set1_player2_2nd_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Converted':
                            set1_player1_break_points_converted, set1_player2_break_points_converted = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Winners':
                            set1_player1_winners, set1_player2_winners = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Unforced Errors':
                            set1_player1_unforced_errors, set1_player2_unforced_errors = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Net Points Won':
                            set1_player1_net_points_won, set1_player2_net_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Points In Row':
                            set1_player1_max_points_in_row, set1_player2_max_points_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Points Won':
                            set1_player1_service_points_won, set1_player2_service_points_won = temp_data[0], temp_data[
                                2]
                        elif temp_data[1].strip() == 'Return Points Won':
                            set1_player1_return_points_won, set1_player2_return_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Points Won':
                            set1_player1_total_points_won, set1_player2_total_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Games In Row':
                            set1_player1_max_games_in_row, set1_player2_max_games_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Games Won':
                            set1_player1_service_games_won, set1_player2_service_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Return Games Won':
                            set1_player1_return_games_won, set1_player2_return_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Games Won':
                            set1_player1_total_games_won, set1_player2_total_games_won = temp_data[0], temp_data[2]

                set_2_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 2"]')
                if len(set_2_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_2_clicks_elm[0])
                    time.sleep(1)
                    set2_elements = driver.find_elements(By.XPATH, '//div[@data-testid="wcl-statistics"]')
                    for set2_element in set2_elements:
                        temp_data = set2_element.text.replace('\n', '#').split('#')
                        if temp_data[1].strip() == 'Aces':
                            set2_player1_aces, set2_player2_aces = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Double Faults':
                            set2_player1_double_faults, set2_player2_double_faults = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Serve Percentage':
                            set2_player1_1st_server_per, set2_player2_1st_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '2nd Serve Points Won':
                            set2_player1_2nd_server_per, set2_player2_2nd_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Saved':
                            set2_player1_break_point_saved, set2_player2_break_point_saved = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Return Points Won':
                            set2_player1_1st_return_points_won, set2_player2_1st_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == '2nd Return Points Won':
                            set2_player1_2nd_return_points_won, set2_player2_2nd_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Converted':
                            set2_player1_break_points_converted, set2_player2_break_points_converted = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Winners':
                            set2_player1_winners, set2_player2_winners = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Unforced Errors':
                            set2_player1_unforced_errors, set2_player2_unforced_errors = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Net Points Won':
                            set2_player1_net_points_won, set2_player2_net_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Points In Row':
                            set2_player1_max_points_in_row, set2_player2_max_points_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Points Won':
                            set2_player1_service_points_won, set2_player2_service_points_won = temp_data[0], temp_data[
                                2]
                        elif temp_data[1].strip() == 'Return Points Won':
                            set2_player1_return_points_won, set2_player2_return_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Points Won':
                            set2_player1_total_points_won, set2_player2_total_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Games In Row':
                            set2_player1_max_games_in_row, set2_player2_max_games_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Games Won':
                            set2_player1_service_games_won, set2_player2_service_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Return Games Won':
                            set2_player1_return_games_won, set2_player2_return_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Games Won':
                            set2_player1_total_games_won, set2_player2_total_games_won = temp_data[0], temp_data[2]

                set_3_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 3"]')
                if len(set_3_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_3_clicks_elm[0])
                    time.sleep(1)
                    set3_elements = driver.find_elements(By.XPATH, '//div[@data-testid="wcl-statistics"]')
                    for set3_element in set3_elements:
                        temp_data = set3_element.text.replace('\n', '#').split('#')
                        if temp_data[1].strip() == 'Aces':
                            set3_player1_aces, set3_player2_aces = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Double Faults':
                            set3_player1_double_faults, set3_player2_double_faults = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Serve Percentage':
                            set3_player1_1st_server_per, set3_player2_1st_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '2nd Serve Points Won':
                            set3_player1_2nd_server_per, set3_player2_2nd_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Saved':
                            set3_player1_break_point_saved, set3_player2_break_point_saved = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Return Points Won':
                            set3_player1_1st_return_points_won, set3_player2_1st_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == '2nd Return Points Won':
                            set3_player1_2nd_return_points_won, set3_player2_2nd_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Converted':
                            set3_player1_break_points_converted, set3_player2_break_points_converted = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Winners':
                            set3_player1_winners, set3_player2_winners = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Unforced Errors':
                            set3_player1_unforced_errors, set3_player2_unforced_errors = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Net Points Won':
                            set3_player1_net_points_won, set3_player2_net_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Points In Row':
                            set3_player1_max_points_in_row, set3_player2_max_points_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Points Won':
                            set3_player1_service_points_won, set3_player2_service_points_won = temp_data[0], temp_data[
                                2]
                        elif temp_data[1].strip() == 'Return Points Won':
                            set3_player1_return_points_won, set3_player2_return_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Points Won':
                            set3_player1_total_points_won, set3_player2_total_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Games In Row':
                            set3_player1_max_games_in_row, set3_player2_max_games_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Games Won':
                            set3_player1_service_games_won, set3_player2_service_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Return Games Won':
                            set3_player1_return_games_won, set3_player2_return_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Games Won':
                            set3_player1_total_games_won, set3_player2_total_games_won = temp_data[0], temp_data[2]

                set_4_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 4"]')
                if len(set_4_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_4_clicks_elm[0])
                    time.sleep(1)
                    set4_elements = driver.find_elements(By.XPATH, '//div[@data-testid="wcl-statistics"]')
                    for set4_element in set4_elements:
                        temp_data = set4_element.text.replace('\n', '#').split('#')
                        if temp_data[1].strip() == 'Aces':
                            set4_player1_aces, set4_player2_aces = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Double Faults':
                            set4_player1_double_faults, set4_player2_double_faults = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Serve Percentage':
                            set4_player1_1st_server_per, set4_player2_1st_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '2nd Serve Points Won':
                            set4_player1_2nd_server_per, set4_player2_2nd_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Saved':
                            set4_player1_break_point_saved, set4_player2_break_point_saved = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Return Points Won':
                            set4_player1_1st_return_points_won, set4_player2_1st_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == '2nd Return Points Won':
                            set4_player1_2nd_return_points_won, set4_player2_2nd_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Converted':
                            set4_player1_break_points_converted, set4_player2_break_points_converted = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Winners':
                            set4_player1_winners, set4_player2_winners = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Unforced Errors':
                            set4_player1_unforced_errors, set4_player2_unforced_errors = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Net Points Won':
                            set4_player1_net_points_won, set4_player2_net_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Points In Row':
                            set4_player1_max_points_in_row, set4_player2_max_points_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Points Won':
                            set4_player1_service_points_won, set4_player2_service_points_won = temp_data[0], temp_data[
                                2]
                        elif temp_data[1].strip() == 'Return Points Won':
                            set4_player1_return_points_won, set4_player2_return_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Points Won':
                            set4_player1_total_points_won, set4_player2_total_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Games In Row':
                            set4_player1_max_games_in_row, set4_player2_max_games_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Games Won':
                            set4_player1_service_games_won, set4_player2_service_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Return Games Won':
                            set4_player1_return_games_won, set4_player2_return_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Games Won':
                            set4_player1_total_games_won, set4_player2_total_games_won = temp_data[0], temp_data[2]

                set_5_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 5"]')
                if len(set_5_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_5_clicks_elm[0])
                    time.sleep(1)
                    set5_elements = driver.find_elements(By.XPATH, '//div[@data-testid="wcl-statistics"]')
                    for set5_element in set5_elements:
                        temp_data = set5_element.text.replace('\n', '#').split('#')
                        if temp_data[1].strip() == 'Aces':
                            set5_player1_aces, set5_player2_aces = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Double Faults':
                            set5_player1_double_faults, set5_player2_double_faults = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Serve Percentage':
                            set5_player1_1st_server_per, set5_player2_1st_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '2nd Serve Points Won':
                            set5_player1_2nd_server_per, set5_player2_2nd_server_per = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Saved':
                            set5_player1_break_point_saved, set5_player2_break_point_saved = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == '1st Return Points Won':
                            set5_player1_1st_return_points_won, set5_player2_1st_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == '2nd Return Points Won':
                            set5_player1_2nd_return_points_won, set5_player2_2nd_return_points_won = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Break Points Converted':
                            set5_player1_break_points_converted, set5_player2_break_points_converted = temp_data[0], \
                                temp_data[2]
                        elif temp_data[1].strip() == 'Winners':
                            set5_player1_winners, set5_player2_winners = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Unforced Errors':
                            set5_player1_unforced_errors, set5_player2_unforced_errors = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Net Points Won':
                            set5_player1_net_points_won, set5_player2_net_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Points In Row':
                            set5_player1_max_points_in_row, set5_player2_max_points_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Points Won':
                            set5_player1_service_points_won, set5_player2_service_points_won = temp_data[0], temp_data[
                                2]
                        elif temp_data[1].strip() == 'Return Points Won':
                            set5_player1_return_points_won, set5_player2_return_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Points Won':
                            set5_player1_total_points_won, set5_player2_total_points_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Max Games In Row':
                            set5_player1_max_games_in_row, set5_player2_max_games_in_row = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Service Games Won':
                            set5_player1_service_games_won, set5_player2_service_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Return Games Won':
                            set5_player1_return_games_won, set5_player2_return_games_won = temp_data[0], temp_data[2]
                        elif temp_data[1].strip() == 'Total Games Won':
                            set5_player1_total_games_won, set5_player2_total_games_won = temp_data[0], temp_data[2]

            # point by point - set 1
            (pbp_set_1_game_1, pbp_set_1_game_2, pbp_set_1_game_3, pbp_set_1_game_4, pbp_set_1_game_5, pbp_set_1_game_6,
             pbp_set_1_game_7, pbp_set_1_game_8, pbp_set_1_game_9, pbp_set_1_game_10, pbp_set_1_game_11,
             pbp_set_1_game_12, pbp_set_1_game_13, pbp_set_1_game_14, pbp_set_1_game_15, pbp_set_1_game_1_serve,
             pbp_set_1_game_2_serve, pbp_set_1_game_3_serve, pbp_set_1_game_4_serve, pbp_set_1_game_5_serve,
             pbp_set_1_game_6_serve, pbp_set_1_game_7_serve, pbp_set_1_game_8_serve, pbp_set_1_game_9_serve,
             pbp_set_1_game_10_serve, pbp_set_1_game_11_serve, pbp_set_1_game_12_serve, pbp_set_1_game_13_serve,
             pbp_set_1_game_14_serve, pbp_set_1_game_15_serve) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '')

            # point by point - set 2
            (pbp_set_2_game_1, pbp_set_2_game_2, pbp_set_2_game_3, pbp_set_2_game_4, pbp_set_2_game_5, pbp_set_2_game_6,
             pbp_set_2_game_7, pbp_set_2_game_8, pbp_set_2_game_9, pbp_set_2_game_10, pbp_set_2_game_11,
             pbp_set_2_game_12, pbp_set_2_game_13, pbp_set_2_game_14, pbp_set_2_game_15, pbp_set_2_game_1_serve,
             pbp_set_2_game_2_serve, pbp_set_2_game_3_serve, pbp_set_2_game_4_serve, pbp_set_2_game_5_serve,
             pbp_set_2_game_6_serve, pbp_set_2_game_7_serve, pbp_set_2_game_8_serve, pbp_set_2_game_9_serve,
             pbp_set_2_game_10_serve, pbp_set_2_game_11_serve, pbp_set_2_game_12_serve, pbp_set_2_game_13_serve,
             pbp_set_2_game_14_serve, pbp_set_2_game_15_serve) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '')

            # point by point - set 3
            (pbp_set_3_game_1, pbp_set_3_game_2, pbp_set_3_game_3, pbp_set_3_game_4, pbp_set_3_game_5, pbp_set_3_game_6,
             pbp_set_3_game_7, pbp_set_3_game_8, pbp_set_3_game_9, pbp_set_3_game_10, pbp_set_3_game_11,
             pbp_set_3_game_12, pbp_set_3_game_13, pbp_set_3_game_14, pbp_set_3_game_15, pbp_set_3_game_1_serve,
             pbp_set_3_game_2_serve, pbp_set_3_game_3_serve, pbp_set_3_game_4_serve, pbp_set_3_game_5_serve,
             pbp_set_3_game_6_serve, pbp_set_3_game_7_serve, pbp_set_3_game_8_serve, pbp_set_3_game_9_serve,
             pbp_set_3_game_10_serve, pbp_set_3_game_11_serve, pbp_set_3_game_12_serve, pbp_set_3_game_13_serve,
             pbp_set_3_game_14_serve, pbp_set_3_game_15_serve) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '')

            # point by point - set 4
            (pbp_set_4_game_1, pbp_set_4_game_2, pbp_set_4_game_3, pbp_set_4_game_4, pbp_set_4_game_5, pbp_set_4_game_6,
             pbp_set_4_game_7, pbp_set_4_game_8, pbp_set_4_game_9, pbp_set_4_game_10, pbp_set_4_game_11,
             pbp_set_4_game_12, pbp_set_4_game_13, pbp_set_4_game_14, pbp_set_4_game_15, pbp_set_4_game_1_serve,
             pbp_set_4_game_2_serve, pbp_set_4_game_3_serve, pbp_set_4_game_4_serve, pbp_set_4_game_5_serve,
             pbp_set_4_game_6_serve, pbp_set_4_game_7_serve, pbp_set_4_game_8_serve, pbp_set_4_game_9_serve,
             pbp_set_4_game_10_serve, pbp_set_4_game_11_serve, pbp_set_4_game_12_serve, pbp_set_4_game_13_serve,
             pbp_set_4_game_14_serve, pbp_set_4_game_15_serve) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '')

            # point by point - set 5
            (pbp_set_5_game_1, pbp_set_5_game_2, pbp_set_5_game_3, pbp_set_5_game_4, pbp_set_5_game_5, pbp_set_5_game_6,
             pbp_set_5_game_7, pbp_set_5_game_8, pbp_set_5_game_9, pbp_set_5_game_10, pbp_set_5_game_11,
             pbp_set_5_game_12, pbp_set_5_game_13, pbp_set_5_game_14, pbp_set_5_game_15, pbp_set_5_game_1_serve,
             pbp_set_5_game_2_serve, pbp_set_5_game_3_serve, pbp_set_5_game_4_serve, pbp_set_5_game_5_serve,
             pbp_set_5_game_6_serve, pbp_set_5_game_7_serve, pbp_set_5_game_8_serve, pbp_set_5_game_9_serve,
             pbp_set_5_game_10_serve, pbp_set_5_game_11_serve, pbp_set_5_game_12_serve, pbp_set_5_game_13_serve,
             pbp_set_5_game_14_serve, pbp_set_5_game_15_serve) = (
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                '', '', '', '')

            point_by_point = driver.find_elements(By.XPATH, "//a[@href='#/match-summary/point-by-point']")
            if len(point_by_point) != 0:

                # set 1
                driver.execute_script("arguments[0].click();", point_by_point[0])
                time.sleep(1)
                set1_games = driver.find_elements(By.XPATH, '//div[@class="matchHistoryRow__fifteens"]')
                for ind, set1_game in enumerate(set1_games):
                    if ind + 1 == 1:
                        pbp_set_1_game_1 = set1_game.text
                        pbp_set_1_game_1_serve = 'Player A'
                    elif ind + 1 == 2:
                        pbp_set_1_game_2 = set1_game.text
                        pbp_set_1_game_2_serve = 'Player B'
                    elif ind + 1 == 3:
                        pbp_set_1_game_3 = set1_game.text
                        pbp_set_1_game_3_serve = 'Player A'
                    elif ind + 1 == 4:
                        pbp_set_1_game_4 = set1_game.text
                        pbp_set_1_game_4_serve = 'Player B'
                    elif ind + 1 == 5:
                        pbp_set_1_game_5 = set1_game.text
                        pbp_set_1_game_5_serve = 'Player A'
                    elif ind + 1 == 6:
                        pbp_set_1_game_6 = set1_game.text
                        pbp_set_1_game_6_serve = 'Player B'
                    elif ind + 1 == 7:
                        pbp_set_1_game_7 = set1_game.text
                        pbp_set_1_game_7_serve = 'Player A'
                    elif ind + 1 == 8:
                        pbp_set_1_game_8 = set1_game.text
                        pbp_set_1_game_8_serve = 'Player B'
                    elif ind + 1 == 9:
                        pbp_set_1_game_9 = set1_game.text
                        pbp_set_1_game_9_serve = 'Player A'
                    elif ind + 1 == 10:
                        pbp_set_1_game_10 = set1_game.text
                        pbp_set_1_game_10_serve = 'Player B'
                    elif ind + 1 == 11:
                        pbp_set_1_game_11 = set1_game.text
                        pbp_set_1_game_11_serve = 'Player A'
                    elif ind + 1 == 12:
                        pbp_set_1_game_12 = set1_game.text
                        pbp_set_1_game_12_serve = 'Player B'
                    elif ind + 1 == 13:
                        pbp_set_1_game_13 = set1_game.text
                        pbp_set_1_game_13_serve = 'Player A'
                    elif ind + 1 == 14:
                        pbp_set_1_game_14 = set1_game.text
                        pbp_set_1_game_14_serve = 'Player B'
                    elif ind + 1 == 15:
                        pbp_set_1_game_15 = set1_game.text
                        pbp_set_1_game_15_serve = 'Player A'

                # set 2
                set_2_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 2"]')
                if len(set_2_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_2_clicks_elm[0])
                    time.sleep(1)
                    set2_games = driver.find_elements(By.XPATH, '//div[@class="matchHistoryRow__fifteens"]')
                    for ind, set2_game in enumerate(set2_games):
                        if ind + 1 == 1:
                            pbp_set_2_game_1 = set2_game.text
                            pbp_set_2_game_1_serve = 'Player A'
                        elif ind + 1 == 2:
                            pbp_set_2_game_2 = set2_game.text
                            pbp_set_2_game_2_serve = 'Player B'
                        elif ind + 1 == 3:
                            pbp_set_2_game_3 = set2_game.text
                            pbp_set_2_game_3_serve = 'Player A'
                        elif ind + 1 == 4:
                            pbp_set_2_game_4 = set2_game.text
                            pbp_set_2_game_4_serve = 'Player B'
                        elif ind + 1 == 5:
                            pbp_set_2_game_5 = set2_game.text
                            pbp_set_2_game_5_serve = 'Player A'
                        elif ind + 1 == 6:
                            pbp_set_2_game_6 = set2_game.text
                            pbp_set_2_game_6_serve = 'Player B'
                        elif ind + 1 == 7:
                            pbp_set_2_game_7 = set2_game.text
                            pbp_set_2_game_7_serve = 'Player A'
                        elif ind + 1 == 8:
                            pbp_set_2_game_8 = set2_game.text
                            pbp_set_2_game_8_serve = 'Player B'
                        elif ind + 1 == 9:
                            pbp_set_2_game_9 = set2_game.text
                            pbp_set_2_game_9_serve = 'Player A'
                        elif ind + 1 == 10:
                            pbp_set_2_game_10 = set2_game.text
                            pbp_set_2_game_10_serve = 'Player B'
                        elif ind + 1 == 11:
                            pbp_set_2_game_11 = set2_game.text
                            pbp_set_2_game_11_serve = 'Player A'
                        elif ind + 1 == 12:
                            pbp_set_2_game_12 = set2_game.text
                            pbp_set_2_game_12_serve = 'Player B'
                        elif ind + 1 == 13:
                            pbp_set_2_game_13 = set2_game.text
                            pbp_set_2_game_13_serve = 'Player A'
                        elif ind + 1 == 14:
                            pbp_set_2_game_14 = set2_game.text
                            pbp_set_2_game_14_serve = 'Player B'
                        elif ind + 1 == 15:
                            pbp_set_2_game_15 = set2_game.text
                            pbp_set_2_game_15_serve = 'Player A'

                # set 3
                set_3_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 3"]')
                if len(set_3_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_3_clicks_elm[0])
                    time.sleep(1)
                    set3_games = driver.find_elements(By.XPATH, '//div[@class="matchHistoryRow__fifteens"]')
                    for ind, set3_game in enumerate(set3_games):
                        if ind + 1 == 1:
                            pbp_set_3_game_1 = set3_game.text
                            pbp_set_3_game_1_serve = 'Player A'
                        elif ind + 1 == 2:
                            pbp_set_3_game_2 = set3_game.text
                            pbp_set_3_game_2_serve = 'Player B'
                        elif ind + 1 == 3:
                            pbp_set_3_game_3 = set3_game.text
                            pbp_set_3_game_3_serve = 'Player A'
                        elif ind + 1 == 4:
                            pbp_set_3_game_4 = set3_game.text
                            pbp_set_3_game_4_serve = 'Player B'
                        elif ind + 1 == 5:
                            pbp_set_3_game_5 = set3_game.text
                            pbp_set_3_game_5_serve = 'Player A'
                        elif ind + 1 == 6:
                            pbp_set_3_game_6 = set3_game.text
                            pbp_set_3_game_6_serve = 'Player B'
                        elif ind + 1 == 7:
                            pbp_set_3_game_7 = set3_game.text
                            pbp_set_3_game_7_serve = 'Player A'
                        elif ind + 1 == 8:
                            pbp_set_3_game_8 = set3_game.text
                            pbp_set_3_game_8_serve = 'Player B'
                        elif ind + 1 == 9:
                            pbp_set_3_game_9 = set3_game.text
                            pbp_set_3_game_9_serve = 'Player A'
                        elif ind + 1 == 10:
                            pbp_set_3_game_10 = set3_game.text
                            pbp_set_3_game_10_serve = 'Player B'
                        elif ind + 1 == 11:
                            pbp_set_3_game_11 = set3_game.text
                            pbp_set_3_game_11_serve = 'Player A'
                        elif ind + 1 == 12:
                            pbp_set_3_game_12 = set3_game.text
                            pbp_set_3_game_12_serve = 'Player B'
                        elif ind + 1 == 13:
                            pbp_set_3_game_13 = set3_game.text
                            pbp_set_3_game_13_serve = 'Player A'
                        elif ind + 1 == 14:
                            pbp_set_3_game_14 = set3_game.text
                            pbp_set_3_game_14_serve = 'Player B'
                        elif ind + 1 == 15:
                            pbp_set_3_game_15 = set3_game.text
                            pbp_set_3_game_15_serve = 'Player A'

                # set 4
                set_4_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 4"]')
                if len(set_4_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_4_clicks_elm[0])
                    time.sleep(1)
                    set4_games = driver.find_elements(By.XPATH, '//div[@class="matchHistoryRow__fifteens"]')
                    for ind, set4_game in enumerate(set4_games):
                        if ind + 1 == 1:
                            pbp_set_4_game_1 = set4_game.text
                            pbp_set_4_game_1_serve = 'Player A'
                        elif ind + 1 == 2:
                            pbp_set_4_game_2 = set4_game.text
                            pbp_set_4_game_2_serve = 'Player B'
                        elif ind + 1 == 3:
                            pbp_set_4_game_3 = set4_game.text
                            pbp_set_4_game_3_serve = 'Player A'
                        elif ind + 1 == 4:
                            pbp_set_4_game_4 = set4_game.text
                            pbp_set_4_game_4_serve = 'Player B'
                        elif ind + 1 == 5:
                            pbp_set_4_game_5 = set4_game.text
                            pbp_set_4_game_5_serve = 'Player A'
                        elif ind + 1 == 6:
                            pbp_set_4_game_6 = set4_game.text
                            pbp_set_4_game_6_serve = 'Player B'
                        elif ind + 1 == 7:
                            pbp_set_4_game_7 = set4_game.text
                            pbp_set_4_game_7_serve = 'Player A'
                        elif ind + 1 == 8:
                            pbp_set_4_game_8 = set4_game.text
                            pbp_set_4_game_8_serve = 'Player B'
                        elif ind + 1 == 9:
                            pbp_set_4_game_9 = set4_game.text
                            pbp_set_4_game_9_serve = 'Player A'
                        elif ind + 1 == 10:
                            pbp_set_4_game_10 = set4_game.text
                            pbp_set_4_game_10_serve = 'Player B'
                        elif ind + 1 == 11:
                            pbp_set_4_game_11 = set4_game.text
                            pbp_set_4_game_11_serve = 'Player A'
                        elif ind + 1 == 12:
                            pbp_set_4_game_12 = set4_game.text
                            pbp_set_4_game_12_serve = 'Player B'
                        elif ind + 1 == 13:
                            pbp_set_4_game_13 = set4_game.text
                            pbp_set_4_game_13_serve = 'Player A'
                        elif ind + 1 == 14:
                            pbp_set_4_game_14 = set4_game.text
                            pbp_set_4_game_14_serve = 'Player B'
                        elif ind + 1 == 15:
                            pbp_set_4_game_15 = set4_game.text
                            pbp_set_4_game_15_serve = 'Player A'

                # set 5
                set_5_clicks_elm = driver.find_elements(By.XPATH, '//a[@title="Set 5"]')
                if len(set_5_clicks_elm) != 0:
                    driver.execute_script("arguments[0].click();", set_5_clicks_elm[0])
                    time.sleep(1)
                    set5_games = driver.find_elements(By.XPATH, '//div[@class="matchHistoryRow__fifteens"]')
                    for ind, set5_game in enumerate(set5_games):
                        if ind + 1 == 1:
                            pbp_set_5_game_1 = set5_game.text
                            pbp_set_5_game_1_serve = 'Player A'
                        elif ind + 1 == 2:
                            pbp_set_5_game_2 = set5_game.text
                            pbp_set_5_game_2_serve = 'Player B'
                        elif ind + 1 == 3:
                            pbp_set_5_game_3 = set5_game.text
                            pbp_set_5_game_3_serve = 'Player A'
                        elif ind + 1 == 4:
                            pbp_set_5_game_4 = set5_game.text
                            pbp_set_5_game_4_serve = 'Player B'
                        elif ind + 1 == 5:
                            pbp_set_5_game_5 = set5_game.text
                            pbp_set_5_game_5_serve = 'Player A'
                        elif ind + 1 == 6:
                            pbp_set_5_game_6 = set5_game.text
                            pbp_set_5_game_6_serve = 'Player B'
                        elif ind + 1 == 7:
                            pbp_set_5_game_7 = set5_game.text
                            pbp_set_5_game_7_serve = 'Player A'
                        elif ind + 1 == 8:
                            pbp_set_5_game_8 = set5_game.text
                            pbp_set_5_game_8_serve = 'Player B'
                        elif ind + 1 == 9:
                            pbp_set_5_game_9 = set5_game.text
                            pbp_set_5_game_9_serve = 'Player A'
                        elif ind + 1 == 10:
                            pbp_set_5_game_10 = set5_game.text
                            pbp_set_5_game_10_serve = 'Player B'
                        elif ind + 1 == 11:
                            pbp_set_5_game_11 = set5_game.text
                            pbp_set_5_game_11_serve = 'Player A'
                        elif ind + 1 == 12:
                            pbp_set_5_game_12 = set5_game.text
                            pbp_set_5_game_12_serve = 'Player B'
                        elif ind + 1 == 13:
                            pbp_set_5_game_13 = set5_game.text
                            pbp_set_5_game_13_serve = 'Player A'
                        elif ind + 1 == 14:
                            pbp_set_5_game_14 = set5_game.text
                            pbp_set_5_game_14_serve = 'Player B'
                        elif ind + 1 == 15:
                            pbp_set_5_game_15 = set5_game.text
                            pbp_set_5_game_15_serve = 'Player A'

            append_to_excel([
                [match_link, tournament_name, surface, stage, tournament_time, player1, player2, rank1, rank2,
                 player1_set_won,
                 player2_set_won, player1_set1_games_won, player2_set1_games_won, player1_set2_games_won,
                 player2_set2_games_won,
                 player1_set3_games_won, player2_set3_games_won, player1_set4_games_won, player2_set4_games_won,
                 player1_set5_games_won, player2_set5_games_won, player1_bet365, player2_bet365, match_player1_aces,
                 match_player2_aces, match_player1_double_faults, match_player2_double_faults,
                 match_player1_1st_server_per,
                 match_player2_1st_server_per, match_player1_2nd_server_per, match_player2_2nd_server_per,
                 match_player1_break_point_saved,
                 match_player2_break_point_saved, match_player1_1st_return_points_won,
                 match_player2_1st_return_points_won,
                 match_player1_2nd_return_points_won, match_player2_2nd_return_points_won,
                 match_player1_break_points_converted,
                 match_player2_break_points_converted, match_player1_winners, match_player2_winners,
                 match_player1_unforced_errors,
                 match_player2_unforced_errors, match_player1_net_points_won, match_player2_net_points_won,
                 match_player1_max_points_in_row,
                 match_player2_max_points_in_row, match_player1_service_points_won, match_player2_service_points_won,
                 match_player1_return_points_won, match_player2_return_points_won, match_player1_total_points_won,
                 match_player2_total_points_won,
                 match_player1_max_games_in_row, match_player2_max_games_in_row, match_player1_service_games_won,
                 match_player2_service_games_won,
                 match_player1_return_games_won, match_player2_return_games_won, match_player1_total_games_won,
                 match_player2_total_games_won, set1_player1_aces, set1_player2_aces, set1_player1_double_faults,
                 set1_player2_double_faults,
                 set1_player1_1st_server_per,
                 set1_player2_1st_server_per, set1_player1_2nd_server_per, set1_player2_2nd_server_per,
                 set1_player1_break_point_saved,
                 set1_player2_break_point_saved, set1_player1_1st_return_points_won, set1_player2_1st_return_points_won,
                 set1_player1_2nd_return_points_won, set1_player2_2nd_return_points_won,
                 set1_player1_break_points_converted,
                 set1_player2_break_points_converted, set1_player1_winners, set1_player2_winners,
                 set1_player1_unforced_errors,
                 set1_player2_unforced_errors, set1_player1_net_points_won, set1_player2_net_points_won,
                 set1_player1_max_points_in_row,
                 set1_player2_max_points_in_row, set1_player1_service_points_won, set1_player2_service_points_won,
                 set1_player1_return_points_won, set1_player2_return_points_won, set1_player1_total_points_won,
                 set1_player2_total_points_won,
                 set1_player1_max_games_in_row, set1_player2_max_games_in_row, set1_player1_service_games_won,
                 set1_player2_service_games_won,
                 set1_player1_return_games_won, set1_player2_return_games_won, set1_player1_total_games_won,
                 set1_player2_total_games_won, set2_player1_aces, set2_player2_aces, set2_player1_double_faults,
                 set2_player2_double_faults,
                 set2_player1_1st_server_per,
                 set2_player2_1st_server_per, set2_player1_2nd_server_per, set2_player2_2nd_server_per,
                 set2_player1_break_point_saved,
                 set2_player2_break_point_saved, set2_player1_1st_return_points_won, set2_player2_1st_return_points_won,
                 set2_player1_2nd_return_points_won, set2_player2_2nd_return_points_won,
                 set2_player1_break_points_converted,
                 set2_player2_break_points_converted, set2_player1_winners, set2_player2_winners,
                 set2_player1_unforced_errors,
                 set2_player2_unforced_errors, set2_player1_net_points_won, set2_player2_net_points_won,
                 set2_player1_max_points_in_row,
                 set2_player2_max_points_in_row, set2_player1_service_points_won, set2_player2_service_points_won,
                 set2_player1_return_points_won, set2_player2_return_points_won, set2_player1_total_points_won,
                 set2_player2_total_points_won,
                 set2_player1_max_games_in_row, set2_player2_max_games_in_row, set2_player1_service_games_won,
                 set2_player2_service_games_won,
                 set2_player1_return_games_won, set2_player2_return_games_won, set2_player1_total_games_won,
                 set2_player2_total_games_won, set3_player1_aces, set3_player2_aces, set3_player1_double_faults,
                 set3_player2_double_faults,
                 set3_player1_1st_server_per,
                 set3_player2_1st_server_per, set3_player1_2nd_server_per, set3_player2_2nd_server_per,
                 set3_player1_break_point_saved,
                 set3_player2_break_point_saved, set3_player1_1st_return_points_won, set3_player2_1st_return_points_won,
                 set3_player1_2nd_return_points_won, set3_player2_2nd_return_points_won,
                 set3_player1_break_points_converted,
                 set3_player2_break_points_converted, set3_player1_winners, set3_player2_winners,
                 set3_player1_unforced_errors,
                 set3_player2_unforced_errors, set3_player1_net_points_won, set3_player2_net_points_won,
                 set3_player1_max_points_in_row,
                 set3_player2_max_points_in_row, set3_player1_service_points_won, set3_player2_service_points_won,
                 set3_player1_return_points_won, set3_player2_return_points_won, set3_player1_total_points_won,
                 set3_player2_total_points_won,
                 set3_player1_max_games_in_row, set3_player2_max_games_in_row, set3_player1_service_games_won,
                 set3_player2_service_games_won,
                 set3_player1_return_games_won, set3_player2_return_games_won, set3_player1_total_games_won,
                 set3_player2_total_games_won, set4_player1_aces, set4_player2_aces, set4_player1_double_faults,
                 set4_player2_double_faults,
                 set4_player1_1st_server_per,
                 set4_player2_1st_server_per, set4_player1_2nd_server_per, set4_player2_2nd_server_per,
                 set4_player1_break_point_saved,
                 set4_player2_break_point_saved, set4_player1_1st_return_points_won, set4_player2_1st_return_points_won,
                 set4_player1_2nd_return_points_won, set4_player2_2nd_return_points_won,
                 set4_player1_break_points_converted,
                 set4_player2_break_points_converted, set4_player1_winners, set4_player2_winners,
                 set4_player1_unforced_errors,
                 set4_player2_unforced_errors, set4_player1_net_points_won, set4_player2_net_points_won,
                 set4_player1_max_points_in_row,
                 set4_player2_max_points_in_row, set4_player1_service_points_won, set4_player2_service_points_won,
                 set4_player1_return_points_won, set4_player2_return_points_won, set4_player1_total_points_won,
                 set4_player2_total_points_won,
                 set4_player1_max_games_in_row, set4_player2_max_games_in_row, set4_player1_service_games_won,
                 set4_player2_service_games_won,
                 set4_player1_return_games_won, set4_player2_return_games_won, set4_player1_total_games_won,
                 set4_player2_total_games_won, set5_player1_aces, set5_player2_aces, set5_player1_double_faults,
                 set5_player2_double_faults,
                 set5_player1_1st_server_per,
                 set5_player2_1st_server_per, set5_player1_2nd_server_per, set5_player2_2nd_server_per,
                 set5_player1_break_point_saved,
                 set5_player2_break_point_saved, set5_player1_1st_return_points_won, set5_player2_1st_return_points_won,
                 set5_player1_2nd_return_points_won, set5_player2_2nd_return_points_won,
                 set5_player1_break_points_converted,
                 set5_player2_break_points_converted, set5_player1_winners, set5_player2_winners,
                 set5_player1_unforced_errors,
                 set5_player2_unforced_errors, set5_player1_net_points_won, set5_player2_net_points_won,
                 set5_player1_max_points_in_row,
                 set5_player2_max_points_in_row, set5_player1_service_points_won, set5_player2_service_points_won,
                 set5_player1_return_points_won, set5_player2_return_points_won, set5_player1_total_points_won,
                 set5_player2_total_points_won,
                 set5_player1_max_games_in_row, set5_player2_max_games_in_row, set5_player1_service_games_won,
                 set5_player2_service_games_won,
                 set5_player1_return_games_won, set5_player2_return_games_won, set5_player1_total_games_won,
                 set5_player2_total_games_won, pbp_set_1_game_1, pbp_set_1_game_2, pbp_set_1_game_3, pbp_set_1_game_4,
                 pbp_set_1_game_5, pbp_set_1_game_6,
                 pbp_set_1_game_7, pbp_set_1_game_8, pbp_set_1_game_9, pbp_set_1_game_10, pbp_set_1_game_11,
                 pbp_set_1_game_12, pbp_set_1_game_13, pbp_set_1_game_14, pbp_set_1_game_15, pbp_set_1_game_1_serve,
                 pbp_set_1_game_2_serve, pbp_set_1_game_3_serve, pbp_set_1_game_4_serve, pbp_set_1_game_5_serve,
                 pbp_set_1_game_6_serve, pbp_set_1_game_7_serve, pbp_set_1_game_8_serve, pbp_set_1_game_9_serve,
                 pbp_set_1_game_10_serve, pbp_set_1_game_11_serve, pbp_set_1_game_12_serve, pbp_set_1_game_13_serve,
                 pbp_set_1_game_14_serve, pbp_set_1_game_15_serve, pbp_set_2_game_1, pbp_set_2_game_2, pbp_set_2_game_3,
                 pbp_set_2_game_4, pbp_set_2_game_5, pbp_set_2_game_6,
                 pbp_set_2_game_7, pbp_set_2_game_8, pbp_set_2_game_9, pbp_set_2_game_10, pbp_set_2_game_11,
                 pbp_set_2_game_12, pbp_set_2_game_13, pbp_set_2_game_14, pbp_set_2_game_15, pbp_set_2_game_1_serve,
                 pbp_set_2_game_2_serve, pbp_set_2_game_3_serve, pbp_set_2_game_4_serve, pbp_set_2_game_5_serve,
                 pbp_set_2_game_6_serve, pbp_set_2_game_7_serve, pbp_set_2_game_8_serve, pbp_set_2_game_9_serve,
                 pbp_set_2_game_10_serve, pbp_set_2_game_11_serve, pbp_set_2_game_12_serve, pbp_set_2_game_13_serve,
                 pbp_set_2_game_14_serve, pbp_set_2_game_15_serve, pbp_set_3_game_1, pbp_set_3_game_2, pbp_set_3_game_3,
                 pbp_set_3_game_4, pbp_set_3_game_5, pbp_set_3_game_6,
                 pbp_set_3_game_7, pbp_set_3_game_8, pbp_set_3_game_9, pbp_set_3_game_10, pbp_set_3_game_11,
                 pbp_set_3_game_12, pbp_set_3_game_13, pbp_set_3_game_14, pbp_set_3_game_15, pbp_set_3_game_1_serve,
                 pbp_set_3_game_2_serve, pbp_set_3_game_3_serve, pbp_set_3_game_4_serve, pbp_set_3_game_5_serve,
                 pbp_set_3_game_6_serve, pbp_set_3_game_7_serve, pbp_set_3_game_8_serve, pbp_set_3_game_9_serve,
                 pbp_set_3_game_10_serve, pbp_set_3_game_11_serve, pbp_set_3_game_12_serve, pbp_set_3_game_13_serve,
                 pbp_set_3_game_14_serve, pbp_set_3_game_15_serve, pbp_set_4_game_1, pbp_set_4_game_2, pbp_set_4_game_3,
                 pbp_set_4_game_4, pbp_set_4_game_5, pbp_set_4_game_6,
                 pbp_set_4_game_7, pbp_set_4_game_8, pbp_set_4_game_9, pbp_set_4_game_10, pbp_set_4_game_11,
                 pbp_set_4_game_12, pbp_set_4_game_13, pbp_set_4_game_14, pbp_set_4_game_15, pbp_set_4_game_1_serve,
                 pbp_set_4_game_2_serve, pbp_set_4_game_3_serve, pbp_set_4_game_4_serve, pbp_set_4_game_5_serve,
                 pbp_set_4_game_6_serve, pbp_set_4_game_7_serve, pbp_set_4_game_8_serve, pbp_set_4_game_9_serve,
                 pbp_set_4_game_10_serve, pbp_set_4_game_11_serve, pbp_set_4_game_12_serve, pbp_set_4_game_13_serve,
                 pbp_set_4_game_14_serve, pbp_set_4_game_15_serve, pbp_set_5_game_1, pbp_set_5_game_2, pbp_set_5_game_3,
                 pbp_set_5_game_4, pbp_set_5_game_5, pbp_set_5_game_6,
                 pbp_set_5_game_7, pbp_set_5_game_8, pbp_set_5_game_9, pbp_set_5_game_10, pbp_set_5_game_11,
                 pbp_set_5_game_12, pbp_set_5_game_13, pbp_set_5_game_14, pbp_set_5_game_15, pbp_set_5_game_1_serve,
                 pbp_set_5_game_2_serve, pbp_set_5_game_3_serve, pbp_set_5_game_4_serve, pbp_set_5_game_5_serve,
                 pbp_set_5_game_6_serve, pbp_set_5_game_7_serve, pbp_set_5_game_8_serve, pbp_set_5_game_9_serve,
                 pbp_set_5_game_10_serve, pbp_set_5_game_11_serve, pbp_set_5_game_12_serve, pbp_set_5_game_13_serve,
                 pbp_set_5_game_14_serve, pbp_set_5_game_15_serve]])

            required_time = "{:.2f} seconds".format(time.time() - start_time)
            print(f'{counter}. Match link = {match_link} -> Time required {required_time}')
            counter += 1


def get_tournaments(driver: webdriver.Chrome):
    driver.get('https://www.tennis24.com/')
    elm = driver.find_element(By.ID, "lmenu_5724")
    driver.execute_script("arguments[0].click();", elm)
    time.sleep(1)
    te = []
    tours = driver.find_elements(By.XPATH, '//a[@class="lmc__templateHref"]')
    for tour in tours:
        te.append(tour.get_attribute('href'))

    print(te)


def scrapper():
    print('=============================================================')
    print('Execution starts!')

    driver = config_driver()
    # get_tournaments(driver)
    create_excel_with_header()
    iterate_tournament(driver)

    print('Execution done!')


if __name__ == '__main__':
    scrapper()
